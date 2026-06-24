from __future__ import annotations

import csv
import math
import sys
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Iterable

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise SystemExit(
        "Не установлена библиотека openpyxl.\n"
        "Запустите install_dependencies.bat, затем повторите запуск."
    ) from exc


OUTPUT_NAME = "Итоговая_таблица_подтверждение_рейсов.xlsx"
COMMENT = (
    "По данным оператора рейс не выполнен, по данным перевозчика "
    "рейс выполнен. Требуется подтверждение."
)

# Номера столбцов Excel (A=1, B=2 и т. д.)
COL_ROUTE = 2       # B
COL_PLAN_TIME = 5   # E
COL_OPERATOR = 12   # L
COL_FACT_TIME = 26  # Z
COL_GRZ = 29        # AC
COL_CARRIER = 30    # AD
COL_FACT_KM = 32    # AF
COL_DATE = 3        # C, запасной столбец
COL_TRANSPORT_DAY = 20  # T, предпочтительный столбец

RESULT_HEADERS = [
    "№",
    "Транспортные сутки",
    "№ маршрута",
    "Плановое время начала рейса",
    "Фактическое время начала рейса",
    "ГРЗ",
    "Фактическая транспортная работа, км",
    "Комментарий",
]


@dataclass
class SourceTable:
    rows: list[list[Any]]
    header_row: int  # индекс в rows, начиная с 0
    source_name: str


def is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    return str(value).strip() == ""


def normalize_text(value: Any) -> str:
    return "" if is_blank(value) else str(value).strip()


def normalize_flag(value: Any) -> int | None:
    """Приводит 0, 1, 0.0, 1.0 и их текстовые варианты к int."""
    if is_blank(value) or isinstance(value, bool):
        return None
    text = str(value).strip().replace(",", ".")
    try:
        number = float(text)
    except (TypeError, ValueError):
        return None
    if number == 0:
        return 0
    if number == 1:
        return 1
    return None


def normalize_number(value: Any) -> float | None:
    if is_blank(value):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = str(value).strip().replace(" ", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def cell(row: list[Any], column_number: int) -> Any:
    index = column_number - 1
    return row[index] if index < len(row) else None


def decode_csv(path: Path) -> tuple[str, str]:
    raw = path.read_bytes()
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            return raw.decode(encoding), encoding
        except UnicodeDecodeError as exc:
            last_error = exc
    raise ValueError(f"Не удалось определить кодировку CSV: {last_error}")


def detect_delimiter(sample: str) -> str:
    try:
        return csv.Sniffer().sniff(sample, delimiters=";,\t|").delimiter
    except csv.Error:
        counts = {delimiter: sample.count(delimiter) for delimiter in (";", ",", "\t", "|")}
        return max(counts, key=counts.get)


def read_csv_table(path: Path) -> list[list[Any]]:
    text, _encoding = decode_csv(path)
    delimiter = detect_delimiter(text[:10000])
    return [list(row) for row in csv.reader(text.splitlines(), delimiter=delimiter)]


def read_xlsx_table(path: Path) -> list[list[Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        return [list(row) for row in worksheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def find_header_row(rows: list[list[Any]]) -> int:
    """
    Ищет строку заголовков. Для исходников заказчика проверяет прежде всего
    ожидаемые подписи в B, L и AD, но допускает небольшие отличия в тексте.
    """
    for index, row in enumerate(rows[:100]):
        route = normalize_text(cell(row, COL_ROUTE)).lower()
        operator = normalize_text(cell(row, COL_OPERATOR)).lower()
        carrier = normalize_text(cell(row, COL_CARRIER)).lower()
        if (
            ("маршрут" in route or route in {"№ маршрута", "номер маршрута"})
            and ("выполн" in operator or "рейс" in operator)
            and ("выполн" in carrier or "рейс" in carrier)
        ):
            return index

    # Более мягкий поиск для файлов без всех подписей.
    for index, row in enumerate(rows[:100]):
        values = " | ".join(normalize_text(value).lower() for value in row)
        if "№ маршрута" in values and "плановое время" in values:
            return index
    raise ValueError(
        "Не найдена строка заголовков. Ожидаются столбцы B «№ маршрута», "
        "L и AD с признаком выполнения рейса."
    )


def read_source(path_string: str) -> SourceTable:
    path = Path(path_string)
    if not path.exists():
        raise FileNotFoundError(f"Файл не найден: {path}")
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        rows = read_xlsx_table(path)
    elif suffix == ".csv":
        rows = read_csv_table(path)
    else:
        raise ValueError("Поддерживаются только файлы .xlsx и .csv")
    if not rows:
        raise ValueError("Исходный файл пуст.")
    return SourceTable(rows=rows, header_row=find_header_row(rows), source_name=path.name)


def validate_template(path_string: str) -> None:
    """Проверяет, что выбранный пример читается и содержит ожидаемую структуру."""
    path = Path(path_string)
    if not path.exists():
        raise FileNotFoundError(f"Файл шаблона не найден: {path}")
    if path.suffix.lower() == ".xlsx":
        rows = read_xlsx_table(path)
    elif path.suffix.lower() == ".csv":
        rows = read_csv_table(path)
    else:
        raise ValueError("Шаблон должен иметь формат .xlsx или .csv")

    joined_rows = [
        " | ".join(normalize_text(value).lower() for value in row)
        for row in rows[:100]
    ]
    required_fragments = ("транспорт", "маршрут", "планов", "грз")
    if not any(all(fragment in row for fragment in required_fragments) for row in joined_rows):
        raise ValueError(
            "В шаблоне не найдена строка с ожидаемыми столбцами "
            "(транспортные сутки, маршрут, плановое время, ГРЗ)."
        )


def choose_transport_day(row: list[Any]) -> Any:
    transport_day = cell(row, COL_TRANSPORT_DAY)
    return transport_day if not is_blank(transport_day) else cell(row, COL_DATE)


def normalize_date_value(value: Any) -> str:
    """
    Возвращает дату строго в виде ДД.ММ.ГГГГ.

    Транспортные сутки сохраняются текстом, чтобы Excel при разных
    региональных настройках не переставлял местами день и месяц.
    """
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    text = normalize_text(value)
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).strftime("%d.%m.%Y")
        except ValueError:
            pass
    return text


def normalize_time_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, time):
        return value
    if isinstance(value, (int, float)) and 0 <= float(value) < 1:
        seconds = round(float(value) * 24 * 60 * 60)
        return time((seconds // 3600) % 24, (seconds % 3600) // 60, seconds % 60)
    text = normalize_text(value)
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(text, fmt).time()
        except ValueError:
            pass
    return text


def row_is_empty(row: list[Any]) -> bool:
    relevant = (
        COL_ROUTE, COL_PLAN_TIME, COL_OPERATOR, COL_FACT_TIME,
        COL_GRZ, COL_CARRIER, COL_FACT_KM, COL_DATE, COL_TRANSPORT_DAY,
    )
    return all(is_blank(cell(row, column)) for column in relevant)


def process_rows(source: SourceTable) -> tuple[list[list[Any]], list[list[Any]], dict[str, Any]]:
    confirmation: list[list[Any]] = []
    errors: list[list[Any]] = []
    counters = {
        "total": 0,
        "auto": 0,
        "confirmation": 0,
        "operator_1_carrier_0": 0,
        "both_0": 0,
        "confirmation_km": 0.0,
    }

    for zero_index, row in enumerate(source.rows[source.header_row + 1:], source.header_row + 1):
        excel_row = zero_index + 1
        if row_is_empty(row):
            continue

        counters["total"] += 1
        route = cell(row, COL_ROUTE)
        plan_time = cell(row, COL_PLAN_TIME)
        operator_raw = cell(row, COL_OPERATOR)
        fact_time = cell(row, COL_FACT_TIME)
        grz = cell(row, COL_GRZ)
        carrier_raw = cell(row, COL_CARRIER)
        km_raw = cell(row, COL_FACT_KM)
        transport_day = choose_transport_day(row)
        operator = normalize_flag(operator_raw)
        carrier = normalize_flag(carrier_raw)

        reasons: list[str] = []
        if is_blank(route):
            reasons.append("Отсутствует номер маршрута")
        if is_blank(plan_time):
            reasons.append("Отсутствует плановое время")
        if is_blank(fact_time):
            reasons.append("Отсутствует фактическое время")
        if is_blank(grz):
            reasons.append("Отсутствует ГРЗ")
        if is_blank(operator_raw):
            reasons.append("Не заполнен признак оператора (L)")
        elif operator is None:
            reasons.append(f"Недопустимое значение оператора (L): {operator_raw}")
        if is_blank(carrier_raw):
            reasons.append("Не заполнен признак перевозчика (AD)")
        elif carrier is None:
            reasons.append(f"Недопустимое значение перевозчика (AD): {carrier_raw}")

        if reasons:
            errors.append([
                excel_row,
                normalize_date_value(transport_day),
                route,
                normalize_time_value(plan_time),
                normalize_time_value(fact_time),
                grz,
                operator_raw,
                carrier_raw,
                "; ".join(reasons),
            ])

        if operator == 1 and carrier == 1:
            counters["auto"] += 1
        elif operator == 0 and carrier == 1:
            counters["confirmation"] += 1
            km = normalize_number(km_raw)
            if km is not None:
                counters["confirmation_km"] += km
            confirmation.append([
                len(confirmation) + 1,
                normalize_date_value(transport_day),
                route,
                normalize_time_value(plan_time),
                normalize_time_value(fact_time),
                grz,
                km if km is not None else normalize_text(km_raw),
                COMMENT,
            ])
        elif operator == 1 and carrier == 0:
            counters["operator_1_carrier_0"] += 1
        elif operator == 0 and carrier == 0:
            counters["both_0"] += 1

    counters["confirmation_km"] = round(counters["confirmation_km"], 2)
    return confirmation, errors, counters


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_GRAY = Side(style="thin", color="B7B7B7")
TABLE_BORDER = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)


def style_header(worksheet, row_number: int, column_count: int) -> None:
    for cell_obj in worksheet[row_number][:column_count]:
        cell_obj.fill = HEADER_FILL
        cell_obj.font = HEADER_FONT
        cell_obj.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_obj.border = TABLE_BORDER
    worksheet.row_dimensions[row_number].height = 42


def style_body(worksheet, min_row: int, max_row: int, max_col: int) -> None:
    for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, max_col=max_col):
        for cell_obj in row:
            cell_obj.border = TABLE_BORDER
            cell_obj.alignment = Alignment(vertical="top", wrap_text=True)


def set_column_widths(worksheet, widths: Iterable[float]) -> None:
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width


def build_workbook(
    confirmation: list[list[Any]],
    errors: list[list[Any]],
    counters: dict[str, Any],
    source_name: str,
    template_name: str,
) -> Workbook:
    workbook = Workbook()

    confirmation_sheet = workbook.active
    confirmation_sheet.title = "На подтверждение"
    confirmation_sheet.append(RESULT_HEADERS)
    for result_row in confirmation:
        confirmation_sheet.append(result_row)
    style_header(confirmation_sheet, 1, len(RESULT_HEADERS))
    if confirmation:
        style_body(confirmation_sheet, 2, len(confirmation) + 1, len(RESULT_HEADERS))
    confirmation_sheet.freeze_panes = "A2"
    confirmation_sheet.auto_filter.ref = f"A1:H{max(1, len(confirmation) + 1)}"
    set_column_widths(confirmation_sheet, (7, 20, 14, 22, 22, 16, 24, 55))
    confirmation_sheet.sheet_view.showGridLines = False
    for row_number in range(2, len(confirmation) + 2):
        confirmation_sheet.cell(row_number, 2).number_format = "@"
        confirmation_sheet.cell(row_number, 4).number_format = "HH:MM:SS"
        confirmation_sheet.cell(row_number, 5).number_format = "HH:MM:SS"
        confirmation_sheet.cell(row_number, 7).number_format = "0.00"

    summary_sheet = workbook.create_sheet("Сводка")
    summary_sheet.append(["Показатель", "Значение"])
    summary_rows = [
        ("Исходный файл", source_name),
        ("Шаблон итоговой таблицы", template_name),
        ("Общее количество строк в исходном файле", counters["total"]),
        ("Рейсы, засчитанные автоматически (оператор 1 / перевозчик 1)", counters["auto"]),
        ("Рейсы, требующие подтверждения (оператор 0 / перевозчик 1)", counters["confirmation"]),
        ("Оператор 1 / перевозчик 0", counters["operator_1_carrier_0"]),
        ("Оператор 0 / перевозчик 0", counters["both_0"]),
        (
            "Сумма фактической транспортной работы по рейсам на подтверждение, км",
            counters["confirmation_km"],
        ),
        ("Количество строк с ошибками проверки", len(errors)),
        ("Дата формирования", datetime.now()),
    ]
    for summary_row in summary_rows:
        summary_sheet.append(summary_row)
    style_header(summary_sheet, 1, 2)
    style_body(summary_sheet, 2, len(summary_rows) + 1, 2)
    summary_sheet.freeze_panes = "A2"
    summary_sheet.auto_filter.ref = f"A1:B{len(summary_rows) + 1}"
    set_column_widths(summary_sheet, (72, 30))
    summary_sheet.cell(9, 2).number_format = "0.00"
    summary_sheet.cell(11, 2).number_format = "DD.MM.YYYY HH:MM:SS"
    summary_sheet.sheet_view.showGridLines = False

    error_sheet = workbook.create_sheet("Ошибки проверки")
    error_headers = [
        "Строка исходного файла",
        "Транспортные сутки",
        "№ маршрута",
        "Плановое время",
        "Фактическое время",
        "ГРЗ",
        "Оператор (L)",
        "Перевозчик (AD)",
        "Ошибки",
    ]
    error_sheet.append(error_headers)
    for error_row in errors:
        error_sheet.append(error_row)
    style_header(error_sheet, 1, len(error_headers))
    if errors:
        style_body(error_sheet, 2, len(errors) + 1, len(error_headers))
    error_sheet.freeze_panes = "A2"
    error_sheet.auto_filter.ref = f"A1:I{max(1, len(errors) + 1)}"
    set_column_widths(error_sheet, (20, 20, 14, 20, 20, 16, 15, 18, 60))
    error_sheet.sheet_view.showGridLines = False
    for row_number in range(2, len(errors) + 2):
        error_sheet.cell(row_number, 2).number_format = "@"
        error_sheet.cell(row_number, 4).number_format = "HH:MM:SS"
        error_sheet.cell(row_number, 5).number_format = "HH:MM:SS"

    return workbook


def generate_report(source_path: str, template_path: str, output_directory: str) -> tuple[Path, dict[str, Any]]:
    source = read_source(source_path)
    validate_template(template_path)
    confirmation, errors, counters = process_rows(source)
    workbook = build_workbook(
        confirmation,
        errors,
        counters,
        source.source_name,
        Path(template_path).name,
    )
    output_dir = Path(output_directory)
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / OUTPUT_NAME
    workbook.save(output_path)
    return output_path, counters


if __name__ == "__main__":
    if len(sys.argv) != 5 or sys.argv[1] != "--cli":
        raise SystemExit(
            "Пользовательская обработка выполняется через веб-приложение.\n"
            "Запустите run.bat и откройте http://127.0.0.1:5000/"
        )
    result_path, result_counters = generate_report(sys.argv[2], sys.argv[3], sys.argv[4])
    print(f"Готово: {result_path}")
    print(f"На подтверждение: {result_counters['confirmation']}")
